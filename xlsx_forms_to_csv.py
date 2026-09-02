#!/usr/bin/env python3
"""
xlsx_forms_to_csv.py
=====================

Reusable converter: pick specific tabs ("forms") out of an Excel workbook
by their real-world form name (the heading shown in cell A1 of each tab,
NOT the short internal tab code like "DM" or "TUN"), export each selected
tab to its own CSV with the content preserved exactly as stored, and
bundle the CSVs into a single downloadable .zip.

Built for CRF / EDC-style exports where every worksheet looks like:

    Row 1        -> form title (e.g. "DEMOGRAPHY")           <- what the
    Row 2        -> human-readable field labels                  user
    Row 3        -> machine field codes                          calls the
    Row 4+       -> one data row per subject/record                "form"
    (tab name)   -> short internal code (e.g. "DM")

This script only reads the workbook (openpyxl, read-only) and only writes
new CSV/zip files -- it never edits the source .xlsx.

A form that's too wide for one Excel tab (e.g. "Study drug
administrations", split into EX / EX1 / EX2 / EX3 for columns 1-15,
16-30, 31-45, 46-55) is detected automatically: typing the plain form
name pulls every one of its continuation tabs in one go. Typing an exact
tab code (e.g. "EX1") still selects just that single tab.

--------------------------------------------------------------------------
WHY THIS EXISTS / HOW IT'S MEANT TO BE REUSED
--------------------------------------------------------------------------
The same study (or a different study using the same EDC export template)
will produce a new workbook every time data is pulled. The tab codes and
tab order can vary between exports, so this script never hardcodes which
tab is which -- it re-discovers every tab's form name (from cell A1) each
time it runs, then matches whatever form names *you* type against those
titles. That means:

  * You can point it at any workbook that follows the same layout and
    it will work without code changes.
  * You (or a non-technical teammate) tell it which forms you need by
    NAME, not by tab code, and can use whatever capitalisation you like
    ("demography", "DEMOGRAPHY", "Demography" all match the same tab).
  * If you're not sure of the exact wording, run with --list first to see
    every tab's real form name and its internal code side by side.

--------------------------------------------------------------------------
USAGE
--------------------------------------------------------------------------
1) See what forms exist in a workbook (no conversion, just inspection):
     python xlsx_forms_to_csv.py study_export.xlsx --list

2) Convert specific forms by name (case-insensitive, order doesn't matter):
     python xlsx_forms_to_csv.py study_export.xlsx --forms \
         "Demography" "reproductive status" "BODY MEASUREMENT" \
         "Biopsy Collection and Tissue Archival" \
         "Inclusion and Exclusion criteria" \
         "RECIST Target Lesions - Screening" \
         "RECIST Non-Target Lesions - Screening" \
         "RECIST New Lesions (at follow up)" \
         "Biochemistry (at follow up)" \
         "Concomitant Medications" \
         "Adverse Event" \
         "Study drug administrations" \
         "Study Conclusion Form"

3) Same thing, but the form names live in a text file (one per line,
   blank lines and lines starting with # are ignored):
     python xlsx_forms_to_csv.py study_export.xlsx --forms-file forms.txt

4) Interactive picker (numbered checklist, works well for a one-off run):
     python xlsx_forms_to_csv.py study_export.xlsx --interactive

5) Convert every tab in the workbook:
     python xlsx_forms_to_csv.py study_export.xlsx --all

6) Multiple workbooks in one go (each workbook's CSVs are kept in their
   own sub-folder inside the zip so names never collide):
     python xlsx_forms_to_csv.py site_a.xlsx site_b.xlsx --forms "Demography"

Output: a single <name>_csv_export_<timestamp>.zip written to the
current folder (override with --output-dir / --output-name).

Only dependency: openpyxl (pip install openpyxl --break-system-packages)
--------------------------------------------------------------------------
"""

import argparse
import csv
import io
import re
import sys
import zipfile
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit(
        "This script needs openpyxl.\n"
        "Install it with:  pip install openpyxl --break-system-packages"
    )


# --------------------------------------------------------------------------
# Matching form names typed by a user against each tab's real A1 title
# --------------------------------------------------------------------------

_SPLIT_WORDS = re.compile(r"\s+(?:and|&)\s+|,", re.IGNORECASE)
_TOKEN = re.compile(r"[a-z0-9]+")
_STOPWORDS = {"and", "or", "the", "a", "an", "of", "to", "at", "in", "on"}
_TRAILING_PAREN = re.compile(r"\s*\([^()]*\)\s*$")


def base_title(title):
    """Strip a trailing '(...)' suffix off a title, repeatedly, e.g.
    'STUDY DRUG ADMINISTRATIONS (01-15)' -> 'STUDY DRUG ADMINISTRATIONS'.
    Used to detect when several tabs are really just numbered
    continuations of the one form (EX / EX1 / EX2 / EX3 all reduce to
    the same base title) so a single plain-language query can pull all
    of them at once."""
    prev, t = None, title.strip()
    while prev != t:
        prev = t
        t = _TRAILING_PAREN.sub("", t).strip()
    return t or title.strip()


def group_sheets_by_base_title(sheets):
    """Group sheets that share the same base_title (see above). A group
    with more than one member represents one logical form split across
    several tabs purely for Excel's column-count limit."""
    groups = {}
    for s in sheets:
        groups.setdefault(base_title(s["title"]), []).append(s)
    return groups


def _stem(tok):
    """Very small naive stemmer: just strips a trailing plural 's' so
    'medications' lines up with 'medication', without touching short
    words or double-s endings like 'status'/'assessment'."""
    if len(tok) > 4 and tok.endswith("s") and not tok.endswith("ss"):
        return tok[:-1]
    return tok


def normalize_tokens(text):
    """Lowercase, split into alnum tokens, drop filler words, and stem
    plurals so 'RECIST 1.1 Target Lesions-Screening' / 'recist target
    lesion screening' / 'Concomitant Medications' all compare fairly."""
    raw = _TOKEN.findall(text.lower())
    return {_stem(t) for t in raw if t not in _STOPWORDS}


def containment_score(query, title):
    """Symmetric similarity between a user-typed phrase and a tab's title.
    Rewards the query being fully found inside the title (title has extra
    boilerplate words) AND the title being fully found inside the query
    (query has extra descriptive words) -- either case scores well."""
    q = normalize_tokens(query)
    t = normalize_tokens(title)
    if not q or not t:
        return 0.0
    inter = q & t
    if not inter:
        return 0.0
    q_in_t = len(inter) / len(q)
    t_in_q = len(inter) / len(t)
    jaccard = len(inter) / len(q | t)
    return 0.4 * q_in_t + 0.4 * t_in_q + 0.2 * jaccard


def _ranked(query_text, sheets):
    return sorted(
        ((containment_score(query_text, s["title"]), s) for s in sheets),
        key=lambda x: x[0],
        reverse=True,
    )


def _try_split(q_clean, sheets, min_part_score=0.4):
    """If the phrase looks like two form names joined by 'and'/'&'/',',
    try matching each half on its own. Returns (sheets, avg_score) or
    None if the split doesn't produce two distinct, confident matches."""
    if not _SPLIT_WORDS.search(q_clean):
        return None
    parts = [p for p in _SPLIT_WORDS.split(q_clean) if p.strip()]
    if len(parts) < 2:
        return None
    matched = []
    scores = []
    for part in parts:
        ranked = _ranked(part, sheets)
        if not ranked or ranked[0][0] < min_part_score:
            return None
        matched.append(ranked[0][1])
        scores.append(ranked[0][0])
    if len({s["code"] for s in matched}) < 2:
        return None  # both halves landed on the same tab, not really "two forms"
    return matched, sum(scores) / len(scores)


def match_one(query, sheets, min_score=0.25, dominant_gap=0.15):
    """Return (matched_sheets, warnings) for one user-typed form name.

    Order of attempts: exact tab-code match, exact title match, then a
    fuzzy content-word match across "candidates" -- where a candidate is
    either a single tab, or (when several tabs are just numbered
    continuations of the same form, e.g. EX / EX1 / EX2 / EX3, detected
    via their shared base_title) the whole group at once. This means a
    plain-language query like "Study drug administrations" matches and
    returns all of its continuation tabs together, while typing an exact
    tab code (e.g. "EX1") always still selects just that one tab.

    If the best candidate is clearly ahead of the runner-up, it's taken
    directly. If the top two scores are close, that's treated as a
    signal worth investigating: first try splitting the phrase as two
    form names joined by 'and'/'&'/',' (e.g. "Biopsy Collection and
    Tissue Archival" -> two tabs); if that doesn't resolve it, take the
    best guess but clearly report the close runner-up and the flag to
    force it instead, since close lexical scores can't always tell two
    similarly-worded forms apart on wording alone (e.g. "Concomitant
    Medications" vs "Concomitant Medication Assessment")."""
    q_clean = query.strip()
    if not q_clean:
        return [], []

    for s in sheets:
        if q_clean.lower() == s["code"].lower():
            return [s], []

    norm_q = " ".join(sorted(normalize_tokens(q_clean)))
    for s in sheets:
        if norm_q == " ".join(sorted(normalize_tokens(s["title"]))):
            return [s], []

    groups = group_sheets_by_base_title(sheets)
    candidates = []  # (score, [sheets], label)
    for key, members in groups.items():
        label = key if len(members) > 1 else members[0]["title"]
        score = containment_score(q_clean, label)
        candidates.append((score, members, label))
    candidates.sort(key=lambda x: x[0], reverse=True)

    if not candidates:
        return [], [f"No good match found for \"{query}\". Run with --list to see available forms."]

    top_score, top_sheets, top_label = candidates[0]
    second_score = candidates[1][0] if len(candidates) > 1 else 0.0
    gap = top_score - second_score

    if top_score < min_score:
        return [], [f"No good match found for \"{query}\". Run with --list to see available forms."]

    if gap >= dominant_gap:
        return top_sheets, []

    split_result = _try_split(q_clean, sheets)
    if split_result:
        matched, _avg = split_result
        return matched, []

    alt_label, alt_sheets = candidates[1][2], candidates[1][1]
    tone = "ambiguous match" if top_score >= 0.4 else "weak match"
    top_codes = "+".join(s["code"] for s in top_sheets)
    alt_codes = "+".join(s["code"] for s in alt_sheets)
    return top_sheets, [
        f"{tone}: \"{query}\" -> \"{top_label}\" ({top_codes}, score {top_score:.2f}), "
        f"but \"{alt_label}\" ({alt_codes}, score {second_score:.2f}) is close behind. "
        f"If that's the one you actually meant, rerun with --forms {alt_codes.replace('+', ' ')} for that entry."
    ]


# --------------------------------------------------------------------------
# Reading the workbook / writing CSVs, preserving content exactly
# --------------------------------------------------------------------------

def discover_sheets(path, title_cell="A1"):
    """Open a workbook and return one dict per worksheet with its tab code,
    its real-world title (from title_cell), and a quick data-row count."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheets = []
    for name in wb.sheetnames:
        ws = wb[name]
        title = None
        try:
            title = ws[title_cell].value
        except Exception:
            title = None
        title = str(title).strip() if title not in (None, "") else name
        sheets.append({"code": name, "title": title, "ws": ws})
    return wb, sheets


def format_cell(value, number_format=None):
    """Return the exact text that should end up in the CSV for one cell.
    Text stays untouched. True numbers that carry a zero-padding display
    format (e.g. '000' / '0000') are zero-padded so on-screen leading
    zeros survive the trip to CSV, since a bare CSV has no concept of
    Excel's display formatting."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    if isinstance(value, (int, float)) and number_format:
        fmt = number_format.strip()
        if re.fullmatch(r"0+", fmt):
            try:
                return str(int(value)).zfill(len(fmt))
            except (ValueError, OverflowError):
                pass
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def sheet_to_csv_bytes(ws):
    """Render one worksheet to CSV bytes (UTF-8 with BOM, so Excel opens
    non-ASCII characters correctly), preserving every row/column exactly
    as stored, including blank spacer rows some forms use."""
    buf = io.StringIO()
    writer = csv.writer(buf, quoting=csv.QUOTE_MINIMAL, lineterminator="\r\n")
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col):
        writer.writerow([format_cell(c.value, c.number_format) for c in row])
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def sanitize_filename(name):
    name = re.sub(r'[\\/:*?"<>|]', "-", name)
    return re.sub(r"\s+", " ", name).strip()


# --------------------------------------------------------------------------
# CLI
# --------------------------------------------------------------------------

def read_forms_file(path):
    forms = []
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if line and not line.startswith("#"):
                forms.append(line)
    return forms


def interactive_pick(sheets):
    print("\nTabs found in this workbook:\n")
    for i, s in enumerate(sheets, 1):
        rows = max(0, (s["ws"].max_row or 0) - 3)
        print(f"  {i:>3}) {s['title']}  [{s['code']}, {rows} data row(s)]")
    print(
        "\nEnter the numbers of the forms you want, comma-separated "
        "(e.g. 1,3,7), or 'all':"
    )
    choice = input("> ").strip().lower()
    if choice == "all":
        return list(sheets)
    picked = []
    for part in choice.split(","):
        part = part.strip()
        if part.isdigit():
            idx = int(part) - 1
            if 0 <= idx < len(sheets):
                picked.append(sheets[idx])
    return picked


def main():
    ap = argparse.ArgumentParser(
        description="Convert selected tabs of an Excel CRF export to CSV, zipped.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    ap.add_argument("workbooks", nargs="+", help="Path(s) to .xlsx file(s)")
    ap.add_argument("--forms", nargs="+", help="Form names to export (case-insensitive)")
    ap.add_argument("--forms-file", help="Text file with one form name per line")
    ap.add_argument("--interactive", action="store_true", help="Pick forms from a numbered list")
    ap.add_argument("--all", action="store_true", help="Export every tab in the workbook")
    ap.add_argument("--list", action="store_true", help="List each tab's form name and code, then exit")
    ap.add_argument("--title-cell", default="A1", help="Cell holding the form title (default A1)")
    ap.add_argument("--output-dir", default=".", help="Where to write the zip (default: current folder)")
    ap.add_argument("--output-name", help="Zip filename (default: auto-generated)")
    args = ap.parse_args()

    requested = []
    if args.forms:
        requested.extend(args.forms)
    if args.forms_file:
        requested.extend(read_forms_file(args.forms_file))

    all_entries = []  # (arcname, bytes)
    multi_source = len(args.workbooks) > 1

    for wb_path in args.workbooks:
        wb_path = Path(wb_path)
        if not wb_path.exists():
            print(f"!! File not found, skipping: {wb_path}", file=sys.stderr)
            continue

        wb, sheets = discover_sheets(wb_path, args.title_cell)

        if args.list:
            print(f"\n=== {wb_path.name} ===")
            for s in sheets:
                rows = max(0, (s["ws"].max_row or 0) - 3)
                print(f"  {s['title']:<55} [{s['code']}]  {rows} data row(s)")
            continue

        if args.interactive:
            chosen = interactive_pick(sheets)
        elif args.all:
            chosen = list(sheets)
        elif requested:
            chosen, seen_codes = [], set()
            for q in requested:
                matches, warnings = match_one(q, sheets)
                for w in warnings:
                    print(f"!! {w}", file=sys.stderr)
                for s in matches:
                    if s["code"] not in seen_codes:
                        chosen.append(s)
                        seen_codes.add(s["code"])
                if matches:
                    names = ", ".join(f"{s['title']} ({s['code']})" for s in matches)
                    print(f"OK  \"{q}\" -> {names}")
        else:
            ap.error("Provide --forms, --forms-file, --interactive, --all, or --list")

        if not args.list:
            print(f"\n{wb_path.name}: exporting {len(chosen)} tab(s)")
            used_names = set()
            for s in chosen:
                csv_bytes = sheet_to_csv_bytes(s["ws"])
                base = sanitize_filename(s["title"])
                fname = base + ".csv"
                if fname.lower() in used_names:
                    fname = f"{base} ({s['code']}).csv"
                used_names.add(fname.lower())
                arcname = f"{sanitize_filename(wb_path.stem)}/{fname}" if multi_source else fname
                all_entries.append((arcname, csv_bytes))
                print(f"   - {arcname}")

        wb.close()

    if args.list:
        return

    if not all_entries:
        print("Nothing was exported (no forms matched, or none selected).", file=sys.stderr)
        sys.exit(1)

    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    zip_name = args.output_name or f"csv_export_{stamp}.zip"
    if not zip_name.lower().endswith(".zip"):
        zip_name += ".zip"
    out_path = out_dir / zip_name

    with zipfile.ZipFile(out_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for arcname, data in all_entries:
            zf.writestr(arcname, data)

    print(f"\nDone. Wrote {len(all_entries)} CSV file(s) to {out_path}")


if __name__ == "__main__":
    main()